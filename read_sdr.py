import openpyxl
import sys
import os

file_path = r'c:\Users\KIIT\Downloads\SDR (1).xlsx'

if not os.path.exists(file_path):
    print(f"Error: File not found at {file_path}")
    sys.exit(1)

try:
    wb = openpyxl.load_workbook(file_path)
    print('Sheet names:', wb.sheetnames)
    print('---')

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f'\n=== Sheet: {sheet_name} ===')
        print(f'Rows: {ws.max_row}, Cols: {ws.max_column}')
        # Print header
        header = [str(cell.value) if cell.value is not None else '' for cell in ws[1]]
        print('Header:', ' | '.join(header))
        
        # Print first 20 rows
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=min(ws.max_row, 20), values_only=False)):
            vals = [str(cell.value) if cell.value is not None else '' for cell in row]
            print(f'Row {i+2}:', ' | '.join(vals))
            
except Exception as e:
    print(f"An error occurred: {e}")
