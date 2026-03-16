import openpyxl
import os
import json

file_path = r'c:\Users\KIIT\Downloads\SDR (1).xlsx'
output_file = r'c:\Users\KIIT\Downloads\Chatbot Ai\tealium_sdr.js'

def generate_script():
    try:
        wb = openpyxl.load_workbook(file_path)
        if 'SDR' in wb.sheetnames:
            ws = wb['SDR']
        else:
            ws = wb.active
            print(f"Warning: 'SDR' sheet not found, using active sheet: {ws.title}")

        print(f"Reading from sheet: {ws.title}")

        # Start of the script
        js_content = ["// Tealium Data Layer Script generated from SDR", "var utag_data = {"]
        
        # Iterate rows starting from 10 (as identified in analysis)
        # Verify if Row 10 is indeed the header or data start. 
        # Previous analysis showed Row 10 had headers like 's.hier1', 'Hierarchy'.
        # Data seems to start after that. Let's look at row 11+
        
        variables_found = 0
        
        for i, row in enumerate(ws.iter_rows(min_row=11, values_only=True)):
            # Column indices (0-based):
            # A (0): ?
            # B (1): Data Point / Description? 
            # C (2): Variable Name (e.g., page_name, site_section)
            
            # Adjust based on previous debug output:
            # Row 10: ['s.hier1', 'Hierarchy', 'hierarchy', 'DOM', ...]
            # So Column C (index 2) is the variable name 'hierarchy'
            # Column B (index 1) is 'Hierarchy' (Description)
            
            var_name = row[2]
            description = row[1]
            
            if var_name:
                var_name = str(var_name).strip()
                description = str(description).strip() if description else ""
                
                # formatting description as comment
                if description:
                    js_content.append(f"    // {description}")
                
                # Add variable with empty string placeholder
                js_content.append(f'    "{var_name}": "",')
                variables_found += 1

        js_content.append("};")
        
        # Join and write to file
        final_script = "\n".join(js_content)
        
        with open(output_file, "w", encoding="utf-8") as f:
            f.write(final_script)
            
        print(f"Successfully generated {output_file} with {variables_found} variables.")

    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    generate_script()
