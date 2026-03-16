import openpyxl
import os
import json

file_path = r'c:\Users\KIIT\Downloads\SDR (1).xlsx'
output_file = r'c:\Users\KIIT\Downloads\Chatbot Ai\sdr_variables.json'

def generate_json_mapping():
    try:
        wb = openpyxl.load_workbook(file_path)
        if 'SDR' in wb.sheetnames:
            ws = wb['SDR']
        else:
            ws = wb.active
            print(f"Warning: 'SDR' sheet not found, using active sheet: {ws.title}")

        print(f"Reading from sheet: {ws.title}")

        sdr_variables = []
        
        # Iterate rows starting from 11 (Data starts here based on previous analysis)
        for i, row in enumerate(ws.iter_rows(min_row=11, values_only=True)):
            # Column B (index 1): Description
            # Column C (index 2): Variable Name
            
            var_name = row[2]
            description = row[1]
            
            if var_name:
                var_name = str(var_name).strip()
                description = str(description).strip() if description else ""
                
                sdr_variables.append({
                    "name": var_name,
                    "description": description
                })

        with open(output_file, "w", encoding="utf-8") as f:
            json.dump(sdr_variables, f, indent=4)
            
        print(f"Successfully generated {output_file} with {len(sdr_variables)} variables.")

    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    generate_json_mapping()
